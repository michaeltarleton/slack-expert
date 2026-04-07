#!/usr/bin/env python3
"""
Convert a downloaded Slack file to markdown.

Usage:
    python convert_to_markdown.py --input /tmp/file.html --output /tmp/out.md
    python convert_to_markdown.py --input /tmp/file.pdf --mimetype application/pdf --output /tmp/out.md

Output (JSON to stdout):
    {
      "ok": true,
      "output_path": "/tmp/out.md",
      "input_type": "text/html",
      "conversion_method": "beautifulsoup4",
      "char_count": 2340,
      "needs_ocr": false
    }

Exit codes:
    0  Success (or needs_ocr=true — partial success)
    1  Missing dependency
    2  Unsupported file type
    3  Conversion error
"""
import argparse
import json
import mimetypes
import os
import subprocess
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# Magic byte → mimetype map
# ---------------------------------------------------------------------------
MAGIC_BYTES: list[tuple[bytes, str]] = [
    (b"%PDF", "application/pdf"),
    (b"PK\x03\x04", "application/zip"),   # docx/xlsx — disambiguate by ext
    (b"\x89PNG", "image/png"),
    (b"\xff\xd8\xff", "image/jpeg"),
    (b"GIF8", "image/gif"),
    (b"RIFF", "image/webp"),               # check bytes 8-12 for WEBP
    (b"<!DOCTYPE", "text/html"),
    (b"<!doctype", "text/html"),
    (b"<html", "text/html"),
    (b"<HTML", "text/html"),
    (b"{", "application/json"),
]

ZIP_EXT_MAP = {
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
}

IMAGE_MIMES = {"image/png", "image/jpeg", "image/gif", "image/webp", "image/bmp", "image/tiff"}


def _fail(code: int, message: str) -> None:
    print(json.dumps({"ok": False, "error": message}), flush=True)
    sys.exit(code)


def detect_type(path: str, hint: str | None = None) -> str:
    """Detect mimetype from magic bytes + extension fallback."""
    if hint and hint not in ("application/octet-stream", ""):
        # Trust the hint unless it's generic
        if not hint.startswith("image/") or Path(path).suffix.lower() in (
            ".png", ".jpg", ".jpeg", ".gif", ".webp"
        ):
            return hint

    try:
        with open(path, "rb") as f:
            header = f.read(32)
    except OSError as e:
        _fail(3, f"Cannot read file: {e}")

    # WEBP check: RIFF....WEBP
    if header[:4] == b"RIFF" and header[8:12] == b"WEBP":
        return "image/webp"

    for magic, mime in MAGIC_BYTES:
        if header[:len(magic)].lower() == magic.lower() or header[:len(magic)] == magic:
            if mime == "application/zip":
                ext = Path(path).suffix.lower().lstrip(".")
                return ZIP_EXT_MAP.get(ext, "application/zip")
            return mime

    # Extension fallback
    mime, _ = mimetypes.guess_type(path)
    return mime or "application/octet-stream"


# ---------------------------------------------------------------------------
# Converters
# ---------------------------------------------------------------------------

def html_to_markdown(path: str) -> str:
    """
    Convert HTML to clean markdown using html2text (primary) with BeautifulSoup
    pre-processing to strip noise and de-duplicate repeated content.
    """
    import re

    try:
        import html2text as _h2t
    except ImportError:
        _fail(1, "html2text not installed: pip install html2text")

    try:
        from bs4 import BeautifulSoup
    except ImportError:
        _fail(1, "beautifulsoup4 not installed: pip install beautifulsoup4")

    with open(path, "r", encoding="utf-8", errors="replace") as f:
        raw = f.read()

    # --- Pre-process with BS4 ---
    soup = BeautifulSoup(raw, "html.parser")

    # Strip noise
    for tag in soup(["script", "style", "head", "meta", "link", "noscript", "img"]):
        tag.decompose()

    # Replace <div> with <p> so html2text treats them as block elements
    for div in soup.find_all("div"):
        div.name = "p"

    cleaned_html = str(soup)

    # --- Convert with html2text ---
    h = _h2t.HTML2Text()
    h.ignore_links = True        # email links are noise
    h.ignore_images = True
    h.body_width = 0             # no line-wrapping
    h.unicode_snob = True
    h.ignore_tables = True       # email HTML uses tables for layout — extract as text
    h.single_line_break = False

    md = h.handle(cleaned_html)

    # Normalize: collapse 3+ blank lines → 2, strip leading/trailing whitespace
    md = re.sub(r"\n{3,}", "\n\n", md).strip()
    return md


def pdf_to_markdown(path: str) -> str:
    try:
        import pymupdf
    except ImportError:
        _fail(1, "PyMuPDF not installed: pip install PyMuPDF")

    doc = pymupdf.open(path)
    pages: list[str] = []
    for i, page in enumerate(doc):
        text = page.get_text("text").strip()
        if text:
            pages.append(f"## Page {i + 1}\n\n{text}")
    doc.close()
    return "\n\n---\n\n".join(pages)


def text_to_markdown(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        content = f.read()
    import re
    content = re.sub(r"\r\n", "\n", content)
    content = re.sub(r"\r", "\n", content)
    return content.strip()


def csv_to_markdown(path: str, max_rows: int = 50) -> str:
    """Use qsv to convert CSV to markdown table (first max_rows rows)."""
    qsv_paths = [
        "/c/Users/mtarleton/scoop/shims/qsv",
        "qsv",
    ]

    qsv_cmd = None
    for qsv in qsv_paths:
        try:
            subprocess.run([qsv, "--version"], capture_output=True, timeout=5)
            qsv_cmd = qsv
            break
        except (FileNotFoundError, subprocess.TimeoutExpired):
            continue

    if not qsv_cmd:
        # Fallback: stdlib csv
        import csv
        lines: list[str] = []
        with open(path, newline="", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                lines.append("| " + " | ".join(row) + " |")
                if i == 0:
                    lines.append("| " + " | ".join(["---"] * len(row)) + " |")
                if i >= max_rows:
                    lines.append(f"\n*... (showing first {max_rows} rows)*")
                    break
        return "\n".join(lines)

    try:
        result = subprocess.run(
            [qsv_cmd, "slice", "--end", str(max_rows), path],
            capture_output=True, text=True, timeout=30
        )
        if result.returncode != 0:
            _fail(3, f"qsv error: {result.stderr}")

        import csv
        import io
        lines = []
        reader = csv.reader(io.StringIO(result.stdout))
        for i, row in enumerate(reader):
            lines.append("| " + " | ".join(row) + " |")
            if i == 0:
                lines.append("| " + " | ".join(["---"] * len(row)) + " |")
        return "\n".join(lines)
    except subprocess.TimeoutExpired:
        _fail(3, "qsv timed out processing CSV")


def xlsx_to_markdown(path: str) -> str:
    try:
        import openpyxl
    except ImportError:
        _fail(1, "openpyxl not installed: pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sections: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_md: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [str(c) if c is not None else "" for c in row]
            rows_md.append("| " + " | ".join(cells) + " |")
            if i == 0:
                rows_md.append("| " + " | ".join(["---"] * len(cells)) + " |")
        if rows_md:
            sections.append(f"## Sheet: {sheet_name}\n\n" + "\n".join(rows_md))
    wb.close()
    return "\n\n".join(sections)


def docx_to_markdown(path: str) -> str:
    try:
        import pymupdf
    except ImportError:
        _fail(1, "PyMuPDF not installed: pip install PyMuPDF")
    doc = pymupdf.open(path)
    pages: list[str] = []
    for i, page in enumerate(doc):
        text = page.get_text("text").strip()
        if text:
            pages.append(f"## Page {i + 1}\n\n{text}")
    doc.close()
    return "\n\n---\n\n".join(pages)


def json_to_markdown(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        content = f.read()
    try:
        data = json.loads(content)
        formatted = json.dumps(data, indent=2, ensure_ascii=False)
    except json.JSONDecodeError:
        formatted = content
    return f"```json\n{formatted}\n```"


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------

MIME_DISPATCH: dict[str, tuple[callable, str]] = {
    "text/html": (html_to_markdown, "beautifulsoup4"),
    "application/pdf": (pdf_to_markdown, "pymupdf"),
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": (xlsx_to_markdown, "openpyxl"),
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": (docx_to_markdown, "pymupdf"),
    "text/plain": (text_to_markdown, "passthrough"),
    "text/csv": (csv_to_markdown, "qsv"),
    "application/json": (json_to_markdown, "stdlib"),
    "application/xml": (text_to_markdown, "passthrough"),
    "text/xml": (text_to_markdown, "passthrough"),
}


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert a Slack file attachment to markdown")
    parser.add_argument("--input", required=True, help="Input file path")
    parser.add_argument("--output", required=True, help="Output .md file path")
    parser.add_argument("--mimetype", help="Hint mimetype (optional)")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        _fail(3, f"Input file not found: {args.input}")

    mimetype = detect_type(args.input, args.mimetype)

    # Image → needs_ocr
    if mimetype in IMAGE_MIMES or mimetype.startswith("image/"):
        # Write empty placeholder so cache path exists
        Path(args.output).write_text("", encoding="utf-8")
        result = {
            "ok": True,
            "output_path": args.output,
            "input_type": mimetype,
            "conversion_method": "ocr_required",
            "char_count": 0,
            "needs_ocr": True,
        }
        print(json.dumps(result), flush=True)
        return

    converter, method = MIME_DISPATCH.get(mimetype, (None, None))

    if converter is None:
        # Try extension-based fallback for text types
        ext = Path(args.input).suffix.lower()
        if ext in (".txt", ".md", ".rst", ".log"):
            converter, method = text_to_markdown, "passthrough"
        elif ext in (".htm", ".html"):
            converter, method = html_to_markdown, "beautifulsoup4"
        else:
            _fail(2, f"Unsupported mimetype: {mimetype}. Paste content manually.")

    try:
        markdown = converter(args.input)
    except SystemExit:
        raise
    except Exception as e:
        _fail(3, f"Conversion failed ({method}): {e}")

    if not markdown.strip():
        markdown = f"*No text content extracted from {Path(args.input).name}*"

    Path(args.output).write_text(markdown, encoding="utf-8")

    result = {
        "ok": True,
        "output_path": args.output,
        "input_type": mimetype,
        "conversion_method": method,
        "char_count": len(markdown),
        "needs_ocr": False,
    }
    print(json.dumps(result), flush=True)


if __name__ == "__main__":
    main()
